#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import re
import sys
import argparse
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def parse_md_file(md_content: str) -> dict:
    """
    解析Markdown文件，提取各级标题和院校招生信息。
    返回结构化的数据字典。
    """
    lines = md_content.splitlines()
    
    # 存储结果
    result = {
        "(一)艺术类本科提前批次": [],
        "(二)艺术类本科批次": [],
        "(三)艺术类高职(专科)批次": []
    }
    
    # 当前状态
    current_sheet = None          # 当前工作表
    current_category = None       # 二级标题（大类）
    current_subcategory = None    # 三级/四级标题（小类）
    
    # 当前院校信息
    current_school = None
    current_group = None
    school_rows = []              # 当前院校的所有行
    
    # 标记是否在院校信息中
    in_school = False
    
    i = 0
    while i < len(lines):
        line = lines[i]
        stripped = line.strip()
        
        # ---- 识别标题 ----
        if stripped.startswith('# '):
            # 一级标题：二、艺术类
            i += 1
            continue
            
        elif stripped.startswith('## '):
            # 二级标题：对应工作表
            title = stripped[3:].strip()
            if '本科提前批次' in title:
                current_sheet = "(一)艺术类本科提前批次"
                current_category = None
                current_subcategory = None
                # 保存之前院校
                if in_school and school_rows:
                    for row in school_rows:
                        result[current_sheet].append(row)
                    school_rows = []
                    in_school = False
                    current_school = None
                    current_group = None
            elif '本科批次' in title:
                current_sheet = "(二)艺术类本科批次"
                current_category = None
                current_subcategory = None
                if in_school and school_rows:
                    for row in school_rows:
                        row["大类"] = current_category if current_category else ""
                        row["小类"] = current_subcategory if current_subcategory else ""
                        result[current_sheet].append(row)
                    school_rows = []
                    in_school = False
                    current_school = None
                    current_group = None
            elif '高职' in title or '专科' in title:
                current_sheet = "(三)艺术类高职(专科)批次"
                current_category = None
                current_subcategory = None
                if in_school and school_rows:
                    for row in school_rows:
                        row["大类"] = current_category if current_category else ""
                        row["小类"] = current_subcategory if current_subcategory else ""
                        result[current_sheet].append(row)
                    school_rows = []
                    in_school = False
                    current_school = None
                    current_group = None
            i += 1
            continue
            
        elif stripped.startswith('### '):
            # 三级标题
            title = stripped[4:].strip()
            
            # 判断是否是分类标题还是院校标题
            # 分类标题格式：(1)xxx 或 1．xxx
            is_category = False
            if re.match(r'^[\(（][0-9]+[\)）]', title) or re.match(r'^[0-9]+[．、.]', title):
                is_category = True
            
            if is_category:
                # 是分类标题
                if current_sheet == "(二)艺术类本科批次":
                    # 三级分类是"1．国家公费师范生"等
                    current_category = title
                    current_subcategory = None
                elif current_sheet == "(三)艺术类高职(专科)批次":
                    # 三级分类是"1．音乐教育类 (声乐主项)"等
                    current_category = title
                    current_subcategory = None
                else:
                    # 提前批次没有分类
                    pass
                # 保存之前院校
                if in_school and school_rows:
                    for row in school_rows:
                        row["大类"] = current_category if current_category else ""
                        row["小类"] = current_subcategory if current_subcategory else ""
                        result[current_sheet].append(row)
                    school_rows = []
                    in_school = False
                    current_school = None
                    current_group = None
            else:
                # 是院校标题
                # 保存之前院校
                if in_school and school_rows:
                    for row in school_rows:
                        row["大类"] = current_category if current_category else ""
                        row["小类"] = current_subcategory if current_subcategory else ""
                        result[current_sheet].append(row)
                    school_rows = []
                    in_school = False
                    current_school = None
                    current_group = None
                
                # 解析院校信息
                school_info = parse_school_title(title)
                current_school = school_info
                in_school = True
                current_group = None
            i += 1
            continue
            
        elif stripped.startswith('#### '):
            # 四级标题
            title = stripped[5:].strip()
            if current_sheet == "(二)艺术类本科批次":
                # 四级标题是 (1)音乐教育类(声乐主项) 等
                current_subcategory = title
            i += 1
            continue
        
        # ---- 处理非标题行 ----
        # 检查是否是院校备注
        if stripped.startswith('**院校备注：') or stripped.startswith('**院校备注:**'):
            note = stripped.replace('**院校备注：', '').replace('**院校备注:**', '').strip()
            # 可能有多个备注行
            if current_school:
                if current_school.get("note"):
                    current_school["note"] += " " + note
                else:
                    current_school["note"] = note
            i += 1
            continue
        
        # 检查是否是专业组行
        if re.match(r'\*\*专业组\s*', stripped):
            # 提取专业组编号
            group_match = re.search(r'\*\*专业组\s*([0-9]+)\*\*', stripped)
            if group_match:
                group_id = group_match.group(1)
                # 提取招生总数
                total_match = re.search(r'(\d+)\s*$', stripped)
                total = total_match.group(1) if total_match else ""
                
                # 提取再选科目
                subject_match = re.search(r'（([^）]+)）', stripped)
                subject = ""
                if subject_match and ('再选' in subject_match.group(1) or '科目' in subject_match.group(1)):
                    subject = subject_match.group(1)
                
                current_group = {
                    "id": group_id,
                    "total": total,
                    "subject": subject,
                    "note": ""
                }
            i += 1
            continue
        
        # 检查是否是专业组备注
        if stripped.startswith('**专业组备注：') or stripped.startswith('**专业组备注:**'):
            if current_group:
                note = stripped.replace('**专业组备注：', '').replace('**专业组备注:**', '').strip()
                current_group["note"] = note
            i += 1
            continue
        
        # 检查是否是专业行
        # 格式：代码    名称    计划    学费
        # 或者：代码    名称    学费 (提前批次)
        # 代码格式：两位字符，可能以0开头或字母开头
        if re.match(r'^[A-Za-z0-9]{2}\s+', stripped):
            parts = stripped.split()
            if len(parts) >= 2:
                code = parts[0]
                # 将代码中的字母O替换为数字0
                code = code.replace('O', '0')
                
                # 提取名称
                name_parts = []
                j = 1
                while j < len(parts):
                    # 如果当前部分是数字、免费、待定，则不是名称部分
                    if re.match(r'^[\d]+$', parts[j]) or parts[j] in ['免费', '待定']:
                        break
                    name_parts.append(parts[j])
                    j += 1
                name = ' '.join(name_parts) if name_parts else ""
                
                # 提取计划和学费
                plan = ""
                fee = ""
                remaining = parts[j:] if j < len(parts) else []
                
                if current_sheet == "(一)艺术类本科提前批次":
                    # 提前批次：只有学费
                    fee_str = ' '.join(remaining)
                    if '免费' in fee_str:
                        fee = '免费'
                    else:
                        fee_match = re.search(r'(\d+)', fee_str)
                        if fee_match:
                            fee = fee_match.group(1)
                else:
                    # 其他批次：计划 学费
                    if remaining:
                        # 第一个可能是计划
                        if re.match(r'^[\d]+$', remaining[0]):
                            plan = remaining[0]
                            if len(remaining) > 1:
                                fee = ' '.join(remaining[1:])
                        else:
                            fee = ' '.join(remaining)
                        
                        # 检查免费或待定
                        if '免费' in fee:
                            fee = '免费'
                        elif '待定' in fee:
                            fee = '待定'
                
                # 提取专业备注（从名称中提取括号内容）
                note = ""
                # 尝试提取专业备注
                if '（' in name and '）' in name:
                    note_match = re.search(r'（([^）]+)）', name)
                    if note_match:
                        note = note_match.group(1)
                # 检查是否有额外的备注在名称后面
                # 有些备注在名称后面以[专业备注:...]形式出现
                if '[' in name and ']' in name:
                    bracket_match = re.search(r'\[([^\]]+)\]', name)
                    if bracket_match:
                        if note:
                            note += " " + bracket_match.group(1)
                        else:
                            note = bracket_match.group(1)
                
                # 构建行数据
                row = {
                    "大类": current_category if current_category else "",
                    "小类": current_subcategory if current_subcategory else "",
                    "院校代码": current_school.get("code", "") if current_school else "",
                    "院校名称": current_school.get("name", "") if current_school else "",
                    "院校备注": current_school.get("note", "") if current_school else "",
                    "所在省": current_school.get("province", "") if current_school else "",
                    "所在市": current_school.get("city", "") if current_school else "",
                    "专业组编号": current_group.get("id", "") if current_group else "",
                    "专业组备注": current_group.get("note", "") if current_group else "",
                    "再选科目": current_group.get("subject", "") if current_group else "",
                    "专业组招生总数": current_group.get("total", "") if current_group else "",
                    "专业代码": code,
                    "专业名称": name,
                    "计划招生人数": plan,
                    "学费(元/年)": fee,
                    "专业备注": note
                }
                school_rows.append(row)
            i += 1
            continue
        
        i += 1
    
    # 处理最后的院校
    if in_school and school_rows:
        for row in school_rows:
            row["大类"] = current_category if current_category else ""
            row["小类"] = current_subcategory if current_subcategory else ""
            if current_sheet:
                result[current_sheet].append(row)
    
    return result


def parse_school_title(title: str) -> dict:
    """解析院校标题，提取院校代码、名称、所在省、市"""
    result = {
        "code": "",
        "name": "",
        "province": "",
        "city": "",
        "note": ""
    }
    
    # 去除加粗标记
    title = title.replace('**', '').strip()
    
    # 提取代码和名称
    match = re.match(r'^([0-9]+)\s+(.+)$', title)
    if match:
        result["code"] = match.group(1).strip()
        full_name = match.group(2).strip()
        
        # 尝试提取省市信息：名称(省市)
        loc_match = re.search(r'^(.+)\(([^)]+)\)$', full_name)
        if loc_match:
            result["name"] = loc_match.group(1).strip()
            location = loc_match.group(2).strip()
            # 判断是省还是市
            if '省' in location or '自治区' in location or '特别行政区' in location:
                result["province"] = location
            elif '市' in location:
                # 直辖市
                if location in ['北京市', '上海市', '天津市', '重庆市']:
                    result["province"] = location
                else:
                    result["city"] = location
            else:
                result["city"] = location
        else:
            result["name"] = full_name
    
    # 如果省为空，尝试从院校名称推断
    if not result["province"] and result["name"]:
        province_map = {
            '北京': '北京市', '上海': '上海市', '天津': '天津市', '重庆': '重庆市',
            '四川': '四川省', '河北': '河北省', '山西': '山西省', '辽宁': '辽宁省',
            '吉林': '吉林省', '黑龙江': '黑龙江省', '江苏': '江苏省', '浙江': '浙江省',
            '安徽': '安徽省', '福建': '福建省', '江西': '江西省', '山东': '山东省',
            '河南': '河南省', '湖北': '湖北省', '湖南': '湖南省', '广东': '广东省',
            '海南': '海南省', '贵州': '贵州省', '云南': '云南省', '陕西': '陕西省',
            '甘肃': '甘肃省', '青海': '青海省', '台湾': '台湾省',
            '内蒙古': '内蒙古自治区', '广西': '广西壮族自治区',
            '西藏': '西藏自治区', '宁夏': '宁夏回族自治区',
            '新疆': '新疆维吾尔自治区', '香港': '香港特别行政区', '澳门': '澳门特别行政区'
        }
        for key, value in province_map.items():
            if result["name"].startswith(key):
                result["province"] = value
                break
    
    return result


def create_excel(data: dict, output_path: str):
    """根据解析的数据创建Excel文件"""
    wb = openpyxl.Workbook()
    # 删除默认sheet
    wb.remove(wb.active)
    
    # 定义列标题
    columns = [
        "大类", "小类", "院校代码", "院校名称", "院校备注",
        "所在省", "所在市", "专业组编号", "专业组备注",
        "再选科目", "专业组招生总数",
        "专业代码", "专业名称", "计划招生人数", "学费(元/年)", "专业备注"
    ]
    
    # 需要自动换行的列
    wrap_columns = ["院校备注", "专业组备注", "专业名称", "专业备注"]
    # 需要整数的列
    int_columns = ["专业组招生总数", "计划招生人数", "学费(元/年)"]
    
    # 需要合并的列
    merge_columns = ["院校备注", "专业组编号", "专业组备注", "再选科目", "专业组招生总数"]
    
    # 创建每个sheet
    sheet_names = ['(一)艺术类本科提前批次', '(二)艺术类本科批次', '(三)艺术类高职(专科)批次']
    
    for sheet_name in sheet_names:
        ws = wb.create_sheet(sheet_name)
        rows_data = data.get(sheet_name, [])
        
        # 写入标题行
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(color="FFFFFF")
            cell.fill = PatternFill(start_color="0072B3", end_color="0072B3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入数据
        for row_idx, row_data in enumerate(rows_data, 2):
            for col_idx, col_name in enumerate(columns, 1):
                value = row_data.get(col_name, "")
                # 处理整数字段
                if col_name in int_columns and value and value not in ["待定", "免费", "(待确定)"]:
                    try:
                        if isinstance(value, str) and value.isdigit():
                            value = int(value)
                    except:
                        pass
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                # 设置对齐
                if col_name in wrap_columns:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="top")
        
        # 设置列宽
        col_widths = {
            "大类": 30, "小类": 30, "院校代码": 12, "院校名称": 25, "院校备注": 40,
            "所在省": 15, "所在市": 15, "专业组编号": 12, "专业组备注": 40,
            "再选科目": 20, "专业组招生总数": 15,
            "专业代码": 12, "专业名称": 40, "计划招生人数": 15, "学费(元/年)": 15, "专业备注": 40
        }
        for col_idx, col_name in enumerate(columns, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 20)
        
        # 冻结首行
        ws.freeze_panes = 'A2'
        
        # 合并相同值的列
        if rows_data:
            for col_name in merge_columns:
                col_idx = columns.index(col_name) + 1
                col_letter = get_column_letter(col_idx)
                start_row = 2
                current_value = None
                
                for row_idx in range(2, len(rows_data) + 2):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if current_value is None:
                        current_value = cell_value
                        start_row = row_idx
                    elif cell_value != current_value:
                        if row_idx - start_row > 1:
                            try:
                                ws.merge_cells(f"{col_letter}{start_row}:{col_letter}{row_idx-1}")
                            except:
                                pass
                        current_value = cell_value
                        start_row = row_idx
                # 处理最后一段
                if start_row <= len(rows_data) + 1 and len(rows_data) + 1 - start_row > 1:
                    try:
                        ws.merge_cells(f"{col_letter}{start_row}:{col_letter}{len(rows_data)+1}")
                    except:
                        pass
    
    # 保存文件
    wb.save(output_path)
    print(f"Excel文件已生成: {output_path}")


def main():
    parser = argparse.ArgumentParser(description='从Markdown招生文件生成Excel')
    parser.add_argument('input_file', help='输入的Markdown文件路径')
    parser.add_argument('output_file', help='输出的Excel文件路径')
    args = parser.parse_args()
    
    # 检查输入文件是否存在
    if not Path(args.input_file).exists():
        print(f"错误: 输入文件 {args.input_file} 不存在")
        sys.exit(1)
    
    # 读取Markdown文件
    with open(args.input_file, 'r', encoding='utf-8') as f:
        md_content = f.read()
    
    # 解析
    print("正在解析Markdown文件...")
    data = parse_md_file(md_content)
    
    # 统计信息
    total_rows = 0
    for sheet_name, rows in data.items():
        total_rows += len(rows)
        print(f"  {sheet_name}: {len(rows)} 行数据")
    print(f"总计: {total_rows} 行数据")
    
    if total_rows == 0:
        print("警告: 没有解析到任何数据，请检查Markdown文件格式")
        return
    
    # 生成Excel
    print("正在生成Excel文件...")
    create_excel(data, args.output_file)
    
    print("完成!")


if __name__ == "__main__":
    main()