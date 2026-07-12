# 由md数据生成excel文件。该代码由claude生成。
# 使用方法：python excelGenerator.py source_json_file dest_md_file
# 例如：python mdGenerator.py /mnt/c/_temp/2026_高考_四川/09.md/physics/2026.sichuan.physics.md /mnt/c/_temp/2026_高考_四川/10.excel/physics/026.sichuan.physics.xlsx

import sys
from parse_final import load_and_preprocess, parse_all

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter

if len(sys.argv) < 3:
    print("请提供至少两个参数")
    print("用法: python mdGenerator.py source_json_file dest_md_file")
    sys.exit(1)

source_md_file = sys.argv[1]
dest_excel_file = sys.argv[2]

OUTPUT = dest_excel_file

HEADERS = [
    '大类', '小类', '院校代码', '院校名称', '院校备注',
    '所在省', '所在市', '专业组编号', '专业组备注', '再选科目',
    '专业组招生总数', '专业代码', '专业名称', '计划招生人数',
    '学费(元/年)', '专业备注',
]

# Columns that support filtering (by position, 1-indexed)
# Per instructions, NOT supporting: 院校备注(5), 专业组编号(8), 专业组备注(9), 再选科目(10), 专业组招生总数(11)
FILTER_COLS = {1, 2, 3, 4, 6, 7, 12, 13, 14, 15}  # 1-indexed

# Columns with wrap text
WRAP_COLS = {5, 9, 13, 16}  # 院校备注, 专业组备注, 专业名称, 专业备注

# Columns with integer data (index in record, 0-based field)
# group_total -> col 11, spec_plan -> col 14, spec_fee -> col 15

BATCH_NAMES = [
    '(一)本科提前批次',
    '(二)本科批次',
    '(三)高职(专科)提前批次',
    '(四)高职(专科)批次',
]

# Style constants
HEADER_FILL = PatternFill('solid', start_color='1F4E79', end_color='1F4E79')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
DATA_FONT = Font(name='Arial', size=10, color='000000')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
DATA_ALIGN_LEFT = Alignment(horizontal='left', vertical='top', wrap_text=False)
DATA_ALIGN_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)


def to_int_val(val):
    """Convert a string value to int if possible, else return original or '(待确定)'."""
    if not val or val.strip() == '':
        return ''
    v = val.strip()
    if v in ('免费', '待定', '(待确定)'):
        return v
    try:
        return int(v)
    except ValueError:
        return '(待确定)'


def build_rows(records, batch_name):
    """Convert records for one batch into list of row tuples."""
    rows = []
    for r in records:
        if r['batch'] != batch_name:
            continue
        
        group_total = to_int_val(r['group_total'])
        spec_plan = to_int_val(r['spec_plan'])
        spec_fee_raw = r['spec_fee']
        if spec_fee_raw in ('免费', '待定'):
            spec_fee = spec_fee_raw
        else:
            spec_fee = to_int_val(spec_fee_raw)
        
        row = (
            r['category'],       # 1 大类
            r['subcategory'],    # 2 小类
            r['school_code'],    # 3 院校代码
            r['school_name'],    # 4 院校名称
            r['school_note'],    # 5 院校备注
            r['province'],       # 6 所在省
            r['city'],           # 7 所在市
            r['group_no'],       # 8 专业组编号
            r['group_note'],     # 9 专业组备注
            r['reselect'],       # 10 再选科目
            group_total,         # 11 专业组招生总数
            r['spec_code'],      # 12 专业代码
            r['spec_name'],      # 13 专业名称
            spec_plan,           # 14 计划招生人数
            spec_fee,            # 15 学费(元/年)
            r['spec_note'],      # 16 专业备注
        )
        rows.append(row)
    return rows


def write_sheet(wb, sheet_name, rows):
    ws = wb.create_sheet(title=sheet_name)
    
    # Write header
    ws.append(HEADERS)
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
    
    # Freeze first row
    ws.freeze_panes = 'A2'
    
    # Write data rows
    for row_data in rows:
        ws.append(list(row_data))
    
    # Apply formatting to data rows
    for row_idx in range(2, len(rows) + 2):
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = DATA_FONT
            if col_idx in WRAP_COLS:
                cell.alignment = DATA_ALIGN_WRAP
            else:
                cell.alignment = DATA_ALIGN_LEFT
    
    # Add auto filter on all columns
    ws.auto_filter.ref = f'A1:{get_column_letter(len(HEADERS))}{len(rows) + 1}'
    
    # Column widths
    col_widths = {
        1: 18,   # 大类
        2: 18,   # 小类
        3: 10,   # 院校代码
        4: 28,   # 院校名称
        5: 40,   # 院校备注
        6: 12,   # 所在省
        7: 12,   # 所在市
        8: 12,   # 专业组编号
        9: 30,   # 专业组备注
        10: 14,  # 再选科目
        11: 14,  # 专业组招生总数
        12: 10,  # 专业代码
        13: 45,  # 专业名称
        14: 14,  # 计划招生人数
        15: 14,  # 学费(元/年)
        16: 35,  # 专业备注
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Row height for header
    ws.row_dimensions[1].height = 30
    
    # Merge cells for non-filter columns where value repeats
    # Columns to merge: 专业组编号(8), 专业组备注(9), 再选科目(10), 专业组招生总数(11)
    # Also: 院校备注(5) when same school
    MERGE_COLS = [5, 8, 9, 10, 11]  # 1-indexed
    
    if len(rows) == 0:
        return
    
    for mc in MERGE_COLS:
        _merge_column(ws, mc, 2, len(rows) + 1)
    
    return ws


def _merge_column(ws, col_idx, start_row, end_row):
    """Merge consecutive cells with the same value in a column."""
    from openpyxl.utils import get_column_letter
    col_letter = get_column_letter(col_idx)
    
    merge_start = start_row
    current_val = ws.cell(row=start_row, column=col_idx).value
    
    for row in range(start_row + 1, end_row + 1):
        cell_val = ws.cell(row=row, column=col_idx).value
        # Check if we should break the merge
        # Break if: value changes, OR if the row starts a new school (col 3 = school_code changed)
        # We track this by checking if school_code (col 3) changed
        school_changed = False
        if col_idx in (8, 9, 10, 11):  # group-level: break on school OR group change
            if row > start_row:
                prev_school = ws.cell(row=row-1, column=3).value
                curr_school = ws.cell(row=row, column=3).value
                if prev_school != curr_school:
                    school_changed = True
                prev_group = ws.cell(row=row-1, column=8).value if col_idx != 8 else None
                curr_group = ws.cell(row=row, column=8).value if col_idx != 8 else None
                if prev_group != curr_group:
                    school_changed = True
        elif col_idx == 5:  # school note: break on school change
            if row > start_row:
                prev_school = ws.cell(row=row-1, column=3).value
                curr_school = ws.cell(row=row, column=3).value
                if prev_school != curr_school:
                    school_changed = True
        
        if cell_val != current_val or school_changed:
            if row - 1 > merge_start:
                ws.merge_cells(
                    start_row=merge_start, start_column=col_idx,
                    end_row=row - 1, end_column=col_idx
                )
                merged_cell = ws.cell(row=merge_start, column=col_idx)
                merged_cell.alignment = Alignment(
                    horizontal='left', vertical='top', wrap_text=(col_idx in WRAP_COLS)
                )
            merge_start = row
            current_val = cell_val
        elif school_changed:
            merge_start = row
            current_val = cell_val
    
    # Merge last group
    if end_row > merge_start:
        ws.merge_cells(
            start_row=merge_start, start_column=col_idx,
            end_row=end_row, end_column=col_idx
        )
        merged_cell = ws.cell(row=merge_start, column=col_idx)
        merged_cell.alignment = Alignment(
            horizontal='left', vertical='top', wrap_text=(col_idx in WRAP_COLS)
        )


def main():
    print("Loading and parsing MD...", file=sys.stderr)
    lines = load_and_preprocess(source_md_file)
    records = parse_all(lines)
    print(f"Total records: {len(records)}", file=sys.stderr)
    
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    for batch_name in BATCH_NAMES:
        print(f"Building sheet: {batch_name}", file=sys.stderr)
        rows = build_rows(records, batch_name)
        print(f"  {len(rows)} rows", file=sys.stderr)
        write_sheet(wb, batch_name, rows)
    
    print(f"Saving to {OUTPUT}...", file=sys.stderr)
    wb.save(OUTPUT)
    print("Done!", file=sys.stderr)


if __name__ == '__main__':
    main()